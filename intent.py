

import os

from openpyxl import Workbook,load_workbook
import threading
from framework.httprequests import httprequest
from framework.framework import gl
from openpyxl.styles import Font, colors, Alignment
import time
from decimal import Decimal

# 测试数据excel路径
excelpath=os.getcwd()+'/test-data/意图测试集 0311.xlsx'  # taibaotestcasereduce.xlsx    faqdata.xlsx
# 测试结果路径
testresult=os.getcwd()+'/test-results/intent-results.xlsx'
url='http://192.144.185.207:8080/v1/openapi'

passed=0
failed=0


class TestsetBean:
    testmodule=''
    testquestion=''
    stdquestion=''
    answer=''
    testresult=''
    remarks=''

if os.path.exists(testresult):
    os.remove(testresult)
# 测试问题对象list
questionsCollection=[]
# 测试结果问题list
questionsCollectionExcel=[]

wb=load_workbook(excelpath)
ws=wb['测试']
excelSize=ws.max_row
for i in range(2, excelSize + 1):
    testsetTemp = TestsetBean();
    testsetTemp.testquestion = ws.cell(i, 1).value
    testsetTemp.stdquestion = ws.cell(i, 2).value
    questionsCollection.append(testsetTemp)
    # print('工作表：'+str(i-1)," >>>>>>      ",testsetTemp.testquestion," | ",testsetTemp.stdquestion)

# 所有测试问的个数
questioncount=len(questionsCollection)

print('\n\n********************************* 测试开始 *********************************')

print('>>>>>> 读取excel测试数据结束 >>>>>>')

rlock=threading.RLock()

class cxThreadWorker(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
    def run(self):
        global questionsCollection
        global rlock
        global passed,failed
        while len(questionsCollection) > 0:
            # print(threading.current_thread().getName())
            # 得到 第一个对象，下标为 0
            standardQBeanTemp = questionsCollection.__getitem__(0)
            answer = standardQBeanTemp.stdquestion
            TestQ = standardQBeanTemp.testquestion

            # 删除掉 下标为 0 第一个对象，须要 加锁
            rlock.acquire()
            try:
                questionsCollection.pop(0)
                print('共：' + str(questioncount) + ' 个, 当前：' + str(questioncount - len(questionsCollection)))
            except Exception:
                break
            finally:
                rlock.release()
            # 头信息
            header = {'Content-Type':'application/json',"cache-control":"no-cache",
                      "Postman-Token":"24786e26-bdb4-47a6-8150-01bdae22728d",'userId': str(gl.getTimestamp()),'appId': 'csbot',
                      }


            # body（JSON） 信息
            bodystringTestQ = '{ "text": "' + str(TestQ) + '"}'


            resposeJsonTestQ = httprequest.sendPostwithHeaders(url, header, bodystringTestQ)

            flag=False

            try:
                module = resposeJsonTestQ['info']['module']
                intent = resposeJsonTestQ['info']['intent']
            except Exception as e:
                flag=True
            if flag==False:
                if gl.repaceSpecialCharactersinString(intent) == gl.repaceSpecialCharactersinString(
                        answer) and module == "task_engine":
                    standardQBeanTemp.testresult = 'pass'
                    # print('---> Pass')
                else:
                    standardQBeanTemp.testresult = 'fail'
                    remarks = '>>>期望返回结果：\n' + answer + '\n>>>实际返回结果：\n' + intent +'\n 出话模块：' + module
                    standardQBeanTemp.remarks = remarks

            else:
                standardQBeanTemp.testresult = 'fail'
                remarks = '>>>接口实际返回结果：\n' + str(resposeJsonTestQ)
                standardQBeanTemp.remarks=remarks

            questionsCollectionExcel.append(standardQBeanTemp)


now=lambda :time.time()
start=now()

subThreads=[]

multiThreadCount=3 # set threads count for running the test
for r in range(multiThreadCount):
    thread = cxThreadWorker()
    thread.start()
    subThreads.append(thread)

for tt in subThreads:
    tt.join()


# ******************************************** 写入结果到excel ********************************************
wb=Workbook()
ws=wb.active
# 表头
ws.column_dimensions['A'].width = 50    # 测试问题
ws.column_dimensions['B'].width = 50    # 标准问题
ws.column_dimensions['C'].width = 8     # 测试结果
ws.column_dimensions['D'].width = 120   # 备注

ws.append(['测试问题','标准问题','测试结果','备注'])
for r in range(len(questionsCollectionExcel)):
    testQuestion=questionsCollectionExcel.__getitem__(r).testquestion
    stdQuestion = questionsCollectionExcel.__getitem__(r).stdquestion
    result=questionsCollectionExcel.__getitem__(r).testresult
    remarks=questionsCollectionExcel.__getitem__(r).remarks
    bold_itatic_24_font_fail = Font(name='Calibri', size=11, italic=False, color=colors.RED, bold=True)
    bold_itatic_24_font_pass = Font(name='Calibri', size=11, italic=False, color=colors.GREEN, bold=True)

    ws.append([testQuestion,stdQuestion, result, remarks])

    alignment = Alignment(wrap_text=True)   # 设置 备注单元格为自动换行

    if result == 'fail':
        ws['C' + str(r + 2)].font = bold_itatic_24_font_fail
        ws['D' + str(r + 2)].alignment = alignment
        failed += 1
    else:
        ws['C' + str(r + 2)].font = bold_itatic_24_font_pass
        ws['D' + str(r + 2)].alignment = alignment
        passed += 1

wb.save(testresult)


print('\n\n>>>>>> 测试结束！ 用时：',str(int(now()-start))+' 秒')
print('>>>>>> 共：',str(passed+failed)+' 个',' 通过率:',str(Decimal((passed/(passed+failed))*100).quantize(Decimal('0.00')))+'%',' 通过：'+str(passed)+' 个，', '失败：'+str(failed)+' 个')
print('\n>>>>>> 测试结果路径：'+testresult)
print('\n\n********************************* 测试结束 *********************************')

